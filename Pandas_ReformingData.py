#!/usr/bin/env python
# coding: utf-8

# In[1]:


import pandas as pd
import openpyxl

wb = openpyxl.load_workbook('D:/THANH PHƯƠNG/University/Python for DA/data.xlsx')

def checkUnnamed(cinDf):
  if 'unnamed' in cinDf.lower():
    return False
  return True

df = pd.read_excel('D:/THANH PHƯƠNG/University/Python for DA/data.xlsx', sheet_name='contacts',
                   skiprows= 6,
                   usecols = checkUnnamed)

# print whole sheet data
print(df)

#convert to dict
rawDict = df.to_dict()
rawDict


# In[2]:


inforDict = {
 0: {'Name': 'Hà Quang Tuấn',
  'Phone': '0937623590',
  'gender': 'Nam',
  'Age': 32,
  'birthPlace': 'Thái Bình',
  'birthDate': '05/01/1990',
  'email': 'tuanhq@gmail.com'},
 1: {'Name': 'Trần Hải Nam',
  'Phone': '0981344528',
  'gender': 'Nam',
  'Age': 31,
  'birthPlace': 'Hải Dương',
  'birthDate': '20/01/1991',
  'email': 'namth@gmail.com'},
 2: {'Name': 'Dương Trung Đức',
  'Phone': '0917613311',
  'gender': 'Nam',
  'birthPlace': 'Hà Nội',
  'birthDate': '25/05/1990',
  'email': 'ducdt@gmail.com'}
}
dict2 = {}
for i in range(len(inforDict)):
    fullName = inforDict[i]['Name']
    phoneNumb = inforDict[i]['Phone']
    gender = inforDict[i]['gender']
    dict2[i] = {
        'ten': fullName, 'sdt': phoneNumb,'gtinh':gender
    }
dict2


# In[3]:


from datetime import date
from datetime import datetime

class convertDict:
  headerList = []
  list1 = []
  formedDict = {}
  def __init__(self, rawDict):
    self.rawDict = rawDict
  def showEle(self):
    for i in self.rawDict:
      print(i)
  def convertList(self):
    for i in range(len(self.rawDict['STT'])):
      fullName = self.rawDict['Họ và tên'][i]
      sdt = '0' + str(self.rawDict['SĐT'][i])
      gender = self.rawDict['Giới tính'][i]
      birthPlace = self.rawDict['Nơi sinh'][i]
      birthDate = self.rawDict['Ngày sinh'][i]
      age = int()
      email = self.rawDict['email'][i]
      convertDict.formedDict[i] = {
          'Name':fullName,
          'Phone':sdt,
          'gender':gender,
          'Age':age,
          'birthPlace':birthPlace,
          'birthDate':birthDate,
          'email': email
      }
    return convertDict.formedDict
  def caculateAge(self):
    for i in range(len(self.rawDict['STT'])):
      #đổi ngày sinh sang dạng str để có strptime đổi dữa liệu dạng từ dạng timestamp sang dạng datetime
      birthDate = datetime.strptime(self.rawDict['Ngày sinh'][i].strftime('%m/%d/%Y'), '%m/%d/%Y')
      today = date.today()
      #incase birth in leap year so the number of days will less than whom not 1 day 
      #So that needed to compare the current day with the birth's
      age = today.year - birthDate.year - ((today.month, today.day) < (birthDate.month, birthDate.day))
      convertDict.formedDict[i]['Age'] = age
    return convertDict.formedDict
  def modifyDate(self):
    for i in range(len(self.rawDict['STT'])):
      #strftime dùng để định dạng lại mẫu ngày tháng mình muốn hiện
      formedDate = datetime.strptime(self.rawDict['Ngày sinh'][i].strftime('%m-%d-%Y'), '%m-%d-%Y').strftime('%d/%m/%Y')
      convertDict.formedDict[i]['birthDate'] = formedDate
  
    return convertDict.formedDict
  @classmethod
  def DicttiList(cls):
    index = 1
    for contact in cls.formedDict.values():
      row = [index] # Danh sách lưu các giá trị cần ghi cho mỗi hàng
      index += 1
      for value in contact.values():
        row.append(value)
      cls.list1.append(row)
    return cls.list1
  # def __repr__(self):
  #   return convertDict.formedDict

Dataa = convertDict(rawDict)
formedData = Dataa.convertList()
addAge = Dataa.caculateAge()
modifyingDate = Dataa.modifyDate()
cmm = Dataa.DicttiList()
cmm
# formedData.showEle()


# In[4]:


import os

p = "D:/THANH PHƯƠNG/University/Python for DA/data.xlsx"
os.path.normpath(p)


class sheetSetting:
    def __init__(self,path,sheetName = str, headersList = list, dataa = list):
        self.path = path
        self.sheetName = sheetName
        self.headersList = headersList
        self.dataa = dataa
        
    #create save 
    def save(self):
      wb.save(self.path) 
    #Create new sheet
    def deleteSheet(self):
      #Delete sheets from previous run
      number_of_sheets = len(wb.worksheets)
      for i in reversed(range(1, number_of_sheets)): # Xóa ngược từ sheet cuối về, tránh bị lỗi
        print(i)
        wb.remove_sheet(wb.worksheets[i])
      sheetSetting.save(self.path)
    
    def adding_newSheet(self):
     #create new sheet
      # newSheet = wb.create_sheet(sheetName)
      # save() 
      #import data
      ###c1: create brand new excel
      with pd.ExcelWriter(self.path, engine="openpyxl", mode="a") as writer:
        newSheet = pd.DataFrame(self.dataa).to_excel(writer,
                                                      sheet_name= self.sheetName,
                                                      header = self.headersList, 
                                                      index=False)
      # newSheet = df.to_excel('/content/ggdrive/MyDrive/Python For Data Analysis/Bài 08/data.xlsx', sheet_name=sheetName)
      #import headers 

      ###c2: append sheet to exsiting excel
      # FilePath = "/content/ggdrive/MyDrive/Python For Data Analysis/Bài 08/data.xlsx"
      # Generating the writer engine

      # Adding the DataFrames to the excel as a new sheet
      # newSheet = pd.DataFrame(dataa).to_excel(writer, sheet_name = sheetName)
      # save()
newSheet = sheetSetting(p,'jupyterNew',['STT','Họ và tên', 'SĐT', 'Giới tính' , 'Tuổi', 'Nơi sinh', 'Ngày sinh','email'], cmm)
saved = newSheet.deleteSheet()
created = newSheet.adding_newSheet()


# In[5]:


#create save 
def save():
  wb.save('D:/THANH PHƯƠNG/University/Python for DA/data.xlsx')

#Create new sheet

def generateNew_sheet(sheetName = str, headersList = str, dataa = list):
  #Delete sheets from previous run
  number_of_sheets = len(wb.worksheets)
  for i in reversed(range(1, number_of_sheets)): # Xóa ngược từ sheet cuối về, tránh bị lỗi
    print(i)
    wb.remove_sheet(wb.worksheets[i])
  save()

 #create new sheet
  # newSheet = wb.create_sheet(sheetName)
  # save() 

  #import data
  ###c1: create brand new excel
  with pd.ExcelWriter('D:/THANH PHƯƠNG/University/Python for DA/data.xlsx', engine="openpyxl", mode="a") as writer:
    newSheet = pd.DataFrame(dataa).to_excel(writer,
                                                  sheet_name=sheetName,
                                                  header= headersList, 
                                                  index=False)
  # newSheet = df.to_excel('/content/ggdrive/MyDrive/Python For Data Analysis/Bài 08/data.xlsx', sheet_name=sheetName)
  #import headers 
  
  ###c2: append sheet to exsiting excel
  # FilePath = "/content/ggdrive/MyDrive/Python For Data Analysis/Bài 08/data.xlsx"
  # Generating the writer engine

  # Adding the DataFrames to the excel as a new sheet
  # newSheet = pd.DataFrame(dataa).to_excel(writer, sheet_name = sheetName)
  # save()





generateNew_sheet('dcmm',['STT','Họ và tên', 'Giới tính' , 'Tuổi', 'SĐT', 'Nơi sinh', 'Ngày sinh','email'], cmm)


# In[ ]:




